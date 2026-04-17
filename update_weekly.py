#!/usr/bin/env python3
"""高配当株週次更新 - yfinanceでデータ取得 → data.json更新 → GitHub Pagesへpush"""
import subprocess, sys
subprocess.run([sys.executable, "-m", "pip", "install", "yfinance", "openpyxl", "-q"], capture_output=True)

import yfinance as yf
import json, math, os
from datetime import date, datetime
import warnings
warnings.filterwarnings('ignore')

json_path = "/Users/shuihiroki/kabuka/data.json"
kabuka_dir = "/Users/shuihiroki/kabuka"

# Google Drive Excelのパス（アクセスできる場合のみ更新）
xl_path = "/Users/shuihiroki/Library/CloudStorage/GoogleDrive-hiroki09090429@gmail.com/マイドライブ/高配当株_価格推移一覧.xlsx"

US_TICKERS = {"VYM","HDV","SPYD"}
stocks = [
    ("9986","蔵王産業","卸売"),("3076","あいHD","卸売"),("8130","サンゲツ","卸売"),
    ("2659","サンエー","小売"),("3333","あさひ","小売"),
    ("4008","住友精化","化学"),("4042","東ソー","化学"),("4097","高圧ガス工業","化学"),
    ("8309","三井住友トラストG","銀行"),("8725","MS&ADインシュアランスHD","保険"),
    ("8593","三菱HCキャピタル","その他金融"),("8584","ジャックス","その他金融"),
    ("6785","鈴木","電気機器"),("7723","愛知時計電機","精密機器"),
    ("3231","野村不動産HD","不動産"),("3003","ヒューリック","不動産"),
    ("2169","CDS","サービス"),("9757","船井総研HD","サービス"),
    ("9769","学究社","サービス"),("4641","アルプス技研","サービス"),
    ("3817","SRAホールディングス","情報・通信"),("3901","マークラインズ","情報・通信"),
    ("4674","クレスコ","情報・通信"),("2003","日東富士製粉","食料品"),
    ("1414","ショーボンドHD","建設"),("1928","積水ハウス","建設"),
    ("6345","アイチコーポレーション","機械"),("9364","上組","倉庫・運輸関連"),
    ("9381","エーアイティー","倉庫・運輸関連"),("5388","クニミネ工業","ガラス・土石製品"),
    ("7989","立川ブラインド工業","金属製品"),("7820","ニホンフラッシュ","その他製品"),
    ("7994","オカムラ","その他製品"),("4540","ツムラ","医薬品"),
    ("1343","NF・J-REIT ETF","J-REIT"),
    ("VYM","VYM（米国高配当）","米国ETF"),("HDV","HDV（米国高配当）","米国ETF"),
    ("SPYD","SPYD（米国高配当）","米国ETF"),
]

tickers_yf = [c+".T" if c not in US_TICKERS else c for c,_,_ in stocks]

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M')} データ取得中...")
all_data = yf.download(tickers_yf, period="20d", interval="1d", auto_adjust=True, progress=False)
closes = all_data["Close"] if "Close" in all_data.columns else all_data.xs("Close", axis=1, level=0)

jp_tickers = [t for t in closes.columns if t not in US_TICKERS]
valid_dates = closes[jp_tickers].dropna(how='all').index[-5:]
date_labels = [d.strftime("%m/%d") for d in valid_dates]

# 既存のlongtermデータを保持
try:
    with open(json_path) as f:
        existing = json.load(f)
    lt_stocks = existing.get("longterm",{}).get("stocks",[])
    longterm_updated_at = existing.get("longterm_updated_at","")
except:
    lt_stocks = []
    longterm_updated_at = ""

weekly_stocks = []
for code, name, sector in stocks:
    key = code+".T" if code not in US_TICKERS else code
    prices = []
    for d in valid_dates:
        try:
            v = closes.loc[d, key]
            prices.append(round(float(v),2) if (v==v and not math.isnan(float(v))) else None)
        except:
            prices.append(None)
    valid = [p for p in prices if p is not None]
    week_pct = round((valid[-1]/valid[0]-1)*100, 2) if len(valid)>=2 else None
    weekly_stocks.append({"code":code,"name":name,"sector":sector,
                          "is_us":code in US_TICKERS,"week_pct":week_pct,"prices":prices})

now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
json_data = {
    "updated_at": now_str,
    "weekly_updated_at": now_str,
    "longterm_updated_at": longterm_updated_at,
    "weekly": {"dates": date_labels, "stocks": weekly_stocks},
    "longterm": {"stocks": lt_stocks}
}
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(json_data, f, ensure_ascii=False, indent=2)
print(f"  JSON更新完了: {date_labels}")

# Excel更新（Google Driveにアクセスできる場合のみ）
xl_updated = False
if os.path.exists(os.path.dirname(xl_path)):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        def pct_color(pct):
            if pct is None: return None
            if pct>=5: return "1A5276"
            if pct>=3: return "2980B9"
            if pct>=1: return "5DADE2"
            if pct>=0: return "D6EAF8"
            if pct>=-1: return "FADBD8"
            if pct>=-3: return "EC7063"
            return "C0392B"
        def fill(h): return PatternFill(fill_type="solid", fgColor=h)
        def bdr():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        wb = load_workbook(xl_path)
        if "📅週次推移" in wb.sheetnames: del wb["📅週次推移"]
        ws = wb.create_sheet("📅週次推移", 0)
        today_str = date.today().strftime("%Y年%m月%d日")
        ws.merge_cells("A1:I1")
        ws["A1"] = f"📅 高配当株 週次推移（更新日：{today_str}）"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = fill("1B2631")
        ws.merge_cells("A2:I2")
        ws["A2"] = "直近5営業日の終値推移｜毎営業日22時に自動更新"
        ws["A2"].fill = fill("2C3E50")
        ws["A2"].font = Font(size=9, color="BDC3C7")
        headers = ["コード","銘柄名","セクター","週間\n騰落率"]+date_labels
        for i,(h,w) in enumerate(zip(headers,[8,22,16,9]+[9]*5),1):
            cell = ws.cell(row=3,column=i,value=h)
            cell.font = Font(bold=True,size=9,color="FFFFFF")
            cell.fill = fill("1C2833")
            cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border = bdr()
            ws.column_dimensions[ws.cell(row=3,column=i).column_letter].width = w
        ws.row_dimensions[3].height = 30
        for row_idx,s in enumerate(weekly_stocks,4):
            row_fill = "1B2631" if row_idx%2==0 else "212F3D"
            for ci,val in enumerate([s["code"],s["name"],s["sector"],None]+s["prices"],1):
                cell = ws.cell(row=row_idx,column=ci,value=val)
                cell.font = Font(size=9,color="ECF0F1")
                cell.fill = fill(row_fill)
                cell.border = bdr()
                cell.alignment = Alignment(horizontal="center",vertical="center")
            if s["week_pct"] is not None:
                pc = ws.cell(row=row_idx,column=4)
                pc.value = s["week_pct"]/100
                pc.number_format = "+0.00%;-0.00%"
                pc.fill = fill(pct_color(s["week_pct"]))
                pc.font = Font(size=9,color="FFFFFF" if abs(s["week_pct"])>=1 else "1B2631")
            for di,p in enumerate(s["prices"]):
                if p is not None:
                    c2 = ws.cell(row=row_idx,column=5+di)
                    c2.value = p
                    c2.number_format = "#,##0.00" if s["is_us"] else "#,##0"
                    if di>0 and s["prices"][di-1]:
                        dp = ((p/s["prices"][di-1])-1)*100
                        c2.fill = fill(pct_color(dp))
                        c2.font = Font(size=9,color="FFFFFF" if abs(dp)>=1 else "ECF0F1")
            ws.row_dimensions[row_idx].height = 18
        wb.save(xl_path)
        xl_updated = True
        print(f"  Excel更新完了")
    except Exception as e:
        print(f"  Excel更新スキップ: {e}")
else:
    print(f"  Google Drive未接続のためExcelスキップ")

# GitHub Pagesへpush（どこからでもアクセス可能にする）
try:
    os.chdir(kabuka_dir)
    # cron環境のKeychain問題対策：トークンファイルから認証情報を取得
    token_file = os.path.expanduser("~/.github_kabuka_token")
    auth_url = "https://github.com/hiroki09090429/kabuka-watch.git"
    if os.path.exists(token_file):
        with open(token_file) as tf:
            token = tf.read().strip()
        if token:
            auth_url = f"https://{token}@github.com/hiroki09090429/kabuka-watch.git"
    subprocess.run(["git","add","data.json"], check=True, capture_output=True)
    result = subprocess.run(["git","diff","--staged","--quiet"], capture_output=True)
    if result.returncode != 0:  # 変更あり
        subprocess.run(["git","commit","-m",f"自動更新: {now_str} {date_labels[0]}〜{date_labels[-1]}"],
                      check=True, capture_output=True)
        subprocess.run(["git","push", auth_url, "main"], check=True, capture_output=True)
        print(f"  GitHub Pages push完了")
    else:
        print(f"  データ変更なし、pushスキップ")
except Exception as e:
    print(f"  GitHub push失敗: {e}")

print(f"✅ {now_str} 完了 (Excel={'更新' if xl_updated else 'スキップ'})")
